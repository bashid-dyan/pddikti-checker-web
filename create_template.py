"""Generate template Excel for PDDIKTI Checker."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Mahasiswa"

headers = ["NO", "NIM", "NAMA MAHASISWA", "PERGURUAN TINGGI"]
header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
center = Alignment(horizontal='center', vertical='center')
left = Alignment(horizontal='left', vertical='center')

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border

# Sample data
samples = [
    [1, "C20118694", "AHMAD SANDI", "UNIVERSITAS TADULAKO"],
    [2, "A1J120014", "HABIBA", "UNIVERSITAS HALU OLEO"],
    [3, "", "", ""],
    [4, "", "", ""],
    [5, "", "", ""],
]

data_font = Font(size=11, name="Arial")
for row_idx, row_data in enumerate(samples, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = center if col_idx == 1 else left

ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 30
ws.column_dimensions['D'].width = 35

ws.freeze_panes = "A2"

wb.save("static/template_data_mahasiswa.xlsx")
print("Template created: static/template_data_mahasiswa.xlsx")
