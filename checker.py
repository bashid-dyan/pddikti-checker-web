"""
PDDIKTI Student Status Checker — Web Version
Refactored dari Pddikti_Checker.py untuk dipakai di Flask web app.
"""

import pandas as pd
import requests
import json
import time
import os
import re
from urllib.parse import quote
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ============================================================
# KONFIGURASI
# ============================================================
API_BASE = "https://pddikti.fastapicloud.dev/api"
API_BACKUP = "https://pddikti.rone.dev/api"
HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
DELAY_SECONDS = 0.3


# ============================================================
# FUNGSI UTILITAS
# ============================================================
def clean_nim(nim):
    return re.sub(r'[\s.\-]', '', str(nim).strip())


def nim_match(nim_excel, nim_pddikti):
    a = clean_nim(nim_excel).upper()
    b = clean_nim(nim_pddikti).upper()
    if not a or not b:
        return False
    return a == b or a in b or b in a


def nama_match(nama_excel, nama_pddikti):
    a = str(nama_excel).strip().upper()
    b = str(nama_pddikti).strip().upper()
    if a == b:
        return True
    if a in b or b in a:
        return True
    words_a = [w for w in a.split() if len(w) >= 3]
    words_b = [w for w in b.split() if len(w) >= 3]
    common = sum(1 for w in words_a if w in words_b)
    return common >= 2 or (common >= 1 and len(words_a) == 1)


# ============================================================
# FUNGSI API
# ============================================================
def api_search(keyword, base_url=API_BASE):
    try:
        url = f"{base_url}/search/mhs/{quote(keyword.strip())}/"
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, list):
                return data
            if isinstance(data, dict) and 'mahasiswa' in data:
                return data['mahasiswa']
        if base_url == API_BASE:
            return api_search(keyword, base_url=API_BACKUP)
    except Exception:
        if base_url == API_BASE:
            return api_search(keyword, base_url=API_BACKUP)
    return []


def api_detail(id_mhs, base_url=API_BASE):
    try:
        url = f"{base_url}/mhs/detail/{quote(id_mhs, safe='=+-_')}/"
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            return resp.json()
        if base_url == API_BASE:
            return api_detail(id_mhs, base_url=API_BACKUP)
    except Exception:
        if base_url == API_BASE:
            return api_detail(id_mhs, base_url=API_BACKUP)
    return None


# ============================================================
# LOGIKA PENCARIAN & PENCOCOKAN
# ============================================================
def find_best_match(results, nim, nama, pt):
    nim_c = clean_nim(nim).upper()

    for r in results:
        if clean_nim(r.get('nim', '')).upper() == nim_c:
            return r, "NIM cocok persis"

    for r in results:
        if nim_match(nim, r.get('nim', '')):
            return r, "NIM cocok partial"

    pt_clean = str(pt).strip().upper()
    for r in results:
        r_pt = str(r.get('nama_pt', '')).strip().upper()
        r_singkat = str(r.get('sinkatan_pt', '')).strip().upper()
        if nama_match(nama, r.get('nama', '')):
            pt_words = [w for w in pt_clean.split() if len(w) > 3]
            for w in pt_words:
                if w in r_pt or w in r_singkat:
                    return r, "Nama+PT cocok"
            if pt_clean in r_pt or r_pt in pt_clean:
                return r, "Nama+PT cocok"
            if r_singkat and (r_singkat in pt_clean or pt_clean in r_singkat):
                return r, "Nama+PT(singkatan) cocok"

    exact = [r for r in results if str(r.get('nama', '')).strip().upper() == str(nama).strip().upper()]
    if len(exact) == 1:
        return exact[0], "Nama unik cocok"

    return None, None


def search_student(nim, nama, pt):
    nim_cleaned = clean_nim(nim)

    results = api_search(nim_cleaned)
    if results:
        match, method = find_best_match(results, nim, nama, pt)
        if match:
            return match, f"Via NIM: {method}"

    time.sleep(DELAY_SECONDS)

    results = api_search(nama)
    if results:
        match, method = find_best_match(results, nim, nama, pt)
        if match:
            return match, f"Via Nama: {method}"

    if nim != nim_cleaned:
        time.sleep(DELAY_SECONDS)
        results = api_search(nim)
        if results:
            match, method = find_best_match(results, nim, nama, pt)
            if match:
                return match, f"Via NIM asli: {method}"

    if results:
        return None, f"Ada {len(results)} hasil tapi NIM/PT tidak cocok"
    return None, "Tidak ditemukan di PDDIKTI"


# ============================================================
# DETEKSI KOLOM
# ============================================================
def detect_columns(df):
    cols = df.columns.tolist()
    col_map = {"nim": None, "nama": None, "pt": None}

    for c in cols:
        c_lower = str(c).lower()
        if "nim" in c_lower or "nip" in c_lower:
            col_map["nim"] = c
        elif "nama" in c_lower and "perguruan" not in c_lower and "pt" not in c_lower and "universitas" not in c_lower:
            col_map["nama"] = c
        elif any(x in c_lower for x in ["perguruan", "universitas", "kampus", "pt"]):
            col_map["pt"] = c

    if not col_map["nim"] and len(cols) > 1:
        col_map["nim"] = cols[1]
    if not col_map["nama"] and len(cols) > 2:
        col_map["nama"] = cols[2]
    if not col_map["pt"] and len(cols) > 4:
        col_map["pt"] = cols[4]

    return col_map


def is_sub_header(row):
    values = [str(v).strip() for v in row.values]
    if all(v == '' or (v.isdigit() and len(v) <= 2) for v in values):
        return True
    if len(values) >= 2 and values[0].isdigit() and values[1].isdigit() and len(values[1]) <= 2:
        return True
    return False


# ============================================================
# FORMAT EXCEL OUTPUT
# ============================================================
def format_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb.active

    total_cols = ws.max_column
    total_rows = ws.max_row
    new_col_start = total_cols - 5

    header_fill_orig = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_fill_new = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Arial")
    data_font = Font(size=10, name="Arial")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for col_idx in range(1, total_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill_orig if col_idx < new_col_start else header_fill_new
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    status_col = new_col_start
    ket_col = total_cols

    for row_idx in range(2, total_rows + 1):
        is_odd = (row_idx % 2 == 1)
        ket_val = str(ws.cell(row=row_idx, column=ket_col).value or "")
        row_color = None
        if "Ditemukan" in ket_val and "Tidak" not in ket_val:
            row_color = green_fill
        elif "Tidak ditemukan" in ket_val:
            row_color = red_fill
        elif "tidak cocok" in ket_val.lower():
            row_color = yellow_fill
        elif "Belum diproses" in ket_val:
            row_color = gray_fill

        for col_idx in range(1, total_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.font = data_font
            if col_idx <= 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            if col_idx >= new_col_start:
                if row_color:
                    cell.fill = row_color
                elif is_odd:
                    cell.fill = stripe_fill
            else:
                if is_odd:
                    cell.fill = stripe_fill

    for col_idx in range(1, new_col_start):
        header_val = str(ws.cell(row=1, column=col_idx).value or "")
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = max(len(header_val) + 4, 12)

    new_widths = [30, 28, 22, 18, 13, 38]
    for i, w in enumerate(new_widths):
        col_letter = ws.cell(row=1, column=new_col_start + i).column_letter
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=total_cols).column_letter}{total_rows}"

    wb.save(filepath)


# ============================================================
# MAIN RUNNER (dengan callback progress)
# ============================================================
def run_checker(input_file, output_file, on_progress=None):
    """
    Jalankan pengecekan PDDIKTI.

    Args:
        input_file:   Path file Excel input
        output_file:  Path file Excel output
        on_progress:  Callback function(current, total, nama, status_text)
    """
    progress_file = os.path.splitext(output_file)[0] + "_progress.json"

    df = pd.read_excel(input_file)

    skip_header_row = False
    if len(df) > 0 and is_sub_header(df.iloc[0]):
        skip_header_row = True

    if skip_header_row:
        df = df.iloc[1:].reset_index(drop=True)

    col_map = detect_columns(df)
    total = len(df)
    results_data = []

    # Load existing progress
    progress = {}
    if os.path.exists(progress_file):
        with open(progress_file, 'r', encoding='utf-8') as f:
            progress = json.load(f)

    for idx, row in df.iterrows():
        nama = str(row[col_map['nama']]).strip()
        nim = str(row[col_map['nim']]).strip()
        pt = str(row[col_map['pt']]).strip() if col_map['pt'] else ""
        key = clean_nim(nim)

        # Cache hit
        if key in progress and "Ditemukan" in progress[key].get("keterangan", ""):
            results_data.append(progress[key])
            if on_progress:
                on_progress(idx + 1, total, nama, f"CACHED: {progress[key].get('status_saat_ini', 'N/A')}")
            continue

        if on_progress:
            on_progress(idx + 1, total, nama, "Mencari...")

        result_info = {
            "status_saat_ini": "",
            "nama_pt_pddikti": "",
            "prodi_pddikti": "",
            "nim_pddikti": "",
            "jenjang": "",
            "keterangan": ""
        }

        matched, method = search_student(nim, nama, pt)

        if matched:
            detail = api_detail(matched['id'])
            if detail:
                result_info["status_saat_ini"] = str(detail.get('status_saat_ini', ''))
                result_info["nama_pt_pddikti"] = str(detail.get('nama_pt', ''))
                result_info["prodi_pddikti"] = str(detail.get('prodi', ''))
                result_info["nim_pddikti"] = str(detail.get('nim', '')).strip()
                result_info["jenjang"] = str(detail.get('jenjang', ''))
                result_info["keterangan"] = f"Ditemukan ({method})"
            else:
                result_info["nama_pt_pddikti"] = matched.get('nama_pt', '')
                result_info["prodi_pddikti"] = matched.get('nama_prodi', '')
                result_info["nim_pddikti"] = matched.get('nim', '')
                result_info["keterangan"] = "Detail tidak tersedia"
        else:
            result_info["keterangan"] = method or "Tidak ditemukan di PDDIKTI"

        progress[key] = result_info
        results_data.append(result_info)

        if on_progress:
            status_text = result_info["status_saat_ini"] or result_info["keterangan"]
            on_progress(idx + 1, total, nama, status_text)

        if (idx + 1) % 10 == 0:
            with open(progress_file, 'w', encoding='utf-8') as f:
                json.dump(progress, f, ensure_ascii=False, indent=2)

        time.sleep(DELAY_SECONDS)

    # Save final progress
    with open(progress_file, 'w', encoding='utf-8') as f:
        json.dump(progress, f, ensure_ascii=False, indent=2)

    # Write output Excel
    df_out = pd.read_excel(input_file)
    offset = 1 if skip_header_row else 0
    padding = [""] * offset

    df_out["STATUS MAHASISWA (PDDIKTI)"] = padding + [r.get("status_saat_ini", "") for r in results_data]
    df_out["PT (PDDIKTI)"] = padding + [r.get("nama_pt_pddikti", "") for r in results_data]
    df_out["PRODI (PDDIKTI)"] = padding + [r.get("prodi_pddikti", "") for r in results_data]
    df_out["NIM (PDDIKTI)"] = padding + [r.get("nim_pddikti", "") for r in results_data]
    df_out["JENJANG"] = padding + [r.get("jenjang", "") for r in results_data]
    df_out["KETERANGAN"] = padding + [r.get("keterangan", "") for r in results_data]

    for col in ["STATUS MAHASISWA (PDDIKTI)", "PT (PDDIKTI)", "PRODI (PDDIKTI)",
                "NIM (PDDIKTI)", "JENJANG", "KETERANGAN"]:
        while len(df_out[col]) < len(df_out):
            df_out[col] = list(df_out[col]) + ["Belum diproses"]

    df_out.to_excel(output_file, index=False)
    format_excel(output_file)

    # Summary
    found = sum(1 for r in results_data if "Ditemukan" in r.get("keterangan", ""))
    not_found = sum(1 for r in results_data if "Tidak ditemukan" in r.get("keterangan", ""))
    no_match = sum(1 for r in results_data if "tidak cocok" in r.get("keterangan", "").lower())

    # Cleanup progress file
    if os.path.exists(progress_file):
        os.remove(progress_file)

    return {
        "total": len(results_data),
        "found": found,
        "not_found": not_found,
        "no_match": no_match
    }
